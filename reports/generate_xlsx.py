"""
Generate Excel annexure for AP Transco DGA Fleet Health Report.
Multi-sheet workbook with formatted tables, color-coded risk/fault, filters, and risk reason column.

Usage: python reports/generate_xlsx.py
Output: reports/AP_Transco_DGA_Fleet_Data_Annexure.xlsx
"""
import os
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from report_helpers import load_data, build_risk_reason, build_short_reason, REPORTS_DIR

# ── Styles ───────────────────────────────────────────────
BLUE = "1F4E79"
HEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", size=9, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=BLUE)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)
RISK_FILLS = {
    "Excellent": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Good": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "Fair": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "Poor": PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid"),
    "Critical": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
}
RISK_FONTS = {
    "Excellent": Font(name="Arial", size=9, color="2E7D32"),
    "Good": Font(name="Arial", size=9, color="1565C0"),
    "Fair": Font(name="Arial", size=9, color="E65100"),
    "Poor": Font(name="Arial", size=9, bold=True, color="C62828"),
    "Critical": Font(name="Arial", size=9, bold=True, color="B71C1C"),
}
FAULT_COLORS = {
    "Normal": "2E7D32", "PD": "0288D1", "D1": "7B1FA2", "D2": "6A1B9A",
    "T1": "F9A825", "T2": "E65100", "T3": "C62828", "DT": "AD1457",
}
RISK_ORDER = {"Critical": 0, "Poor": 1, "Fair": 2, "Good": 3, "Excellent": 4}


def setup_header(ws, headers, widths):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def write_cell(ws, row, col, value, font=DATA_FONT, align=LEFT, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    c.border = THIN_BORDER
    if fill:
        c.fill = fill
    return c


def main():
    data = load_data()
    transformers = data["transformers"]
    substations = data["substations"]
    models = data["models"]

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════
    # Sheet 1: All Transformers (with Reason for Risk Rating)
    # ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "All Transformers"

    headers1 = [
        "Equipment No", "Substation", "Make", "Voltage (kV)", "Capacity (MVA)",
        "YOM", "Age (yrs)", "Fault Type", "Confidence (%)",
        "CHI Score", "Risk Level", "DGAF",
        "Reason for Risk Rating",
        "H2", "CH4", "C2H6", "C2H4", "C2H2", "CO", "CO2", "TDCG",
        "BDV", "Moisture", "Tan Delta", "Acidity",
        "Last Sampled", "Samples",
    ]
    widths1 = [14, 35, 15, 8, 10, 7, 7, 9, 8, 8, 10, 7,
               55,
               8, 8, 8, 8, 8, 8, 8, 8,
               8, 10, 8, 8,
               13, 7]
    setup_header(ws1, headers1, widths1)

    sorted_trs = sorted(transformers, key=lambda t: (RISK_ORDER.get(t.get("risk_level"), 5), t.get("chi") or 999))

    for i, t in enumerate(sorted_trs, 2):
        gases = t.get("gases", {})
        oil = t.get("oil_quality", {})
        risk = t.get("risk_level") or ""
        reason = build_risk_reason(t)

        row_data = [
            t.get("equipment_no"),
            t.get("substation_name") or t.get("substation_id") or "",
            t.get("make") or "",
            t.get("voltage_class"),
            t.get("capacity_mva"),
            t.get("yom"),
            round(t["age_years"], 1) if t.get("age_years") is not None else None,
            t.get("fault_label") or "",
            round(t["confidence"] * 100, 0) if t.get("confidence") is not None else None,
            round(t["chi"], 1) if t.get("chi") is not None else None,
            risk,
            round(t["dgaf"], 1) if t.get("dgaf") is not None else None,
            reason,
            gases.get("H2"), gases.get("CH4"), gases.get("C2H6"),
            gases.get("C2H4"), gases.get("C2H2"), gases.get("CO"), gases.get("CO2"),
            t.get("tdcg"),
            oil.get("BDV"), oil.get("MOISTURE"), oil.get("TAN_DELTA"), oil.get("ACIDITY"),
            t.get("latest_sample_date") or "",
            t.get("sample_count"),
        ]

        for col, val in enumerate(row_data, 1):
            c = write_cell(ws1, i, col, val)
            if col == 11:  # Risk
                c.font = RISK_FONTS.get(risk, DATA_FONT)
                if risk in RISK_FILLS:
                    c.fill = RISK_FILLS[risk]
            elif col == 10 and val is not None and val < 40:
                c.font = Font(name="Arial", size=9, bold=True, color="C62828")
            elif col == 8 and val and val != "Normal":
                c.font = Font(name="Arial", size=9, bold=True, color=FAULT_COLORS.get(val, "333333"))
            elif col == 13:  # Reason
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if risk in ("Poor", "Critical"):
                    c.font = Font(name="Arial", size=9, color="C62828")
                elif risk == "Fair":
                    c.font = Font(name="Arial", size=9, color="E65100")
            if isinstance(val, (int, float)) and val is not None and col != 13:
                c.alignment = RIGHT
                if isinstance(val, float):
                    c.number_format = "0.0"

    print(f"  Sheet 1: {len(sorted_trs)} transformers")

    # ═══════════════════════════════════════════════════════
    # Sheet 2: Substations
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Substations")
    headers2 = ["Substation Name", "ID", "Voltage (kV)", "Transformers",
                "Avg CHI", "Worst Risk", "Last Sampled",
                "Excellent", "Good", "Fair", "Poor", "Critical"]
    widths2 = [40, 22, 8, 10, 8, 10, 13, 8, 8, 8, 8, 8]
    setup_header(ws2, headers2, widths2)

    sorted_subs = sorted(substations, key=lambda s: (RISK_ORDER.get(s.get("worst_risk"), 5), s.get("avg_chi") or 999))
    for i, s in enumerate(sorted_subs, 2):
        rd = s.get("risk_distribution", {})
        worst = s.get("worst_risk", "")
        row = [
            s.get("name") or s.get("id"), s.get("id"), s.get("voltage_class"),
            s.get("transformer_count"),
            round(s["avg_chi"], 1) if s.get("avg_chi") is not None else None,
            worst, s.get("latest_sample_date") or "",
            rd.get("Excellent", 0), rd.get("Good", 0), rd.get("Fair", 0),
            rd.get("Poor", 0), rd.get("Critical", 0),
        ]
        for col, val in enumerate(row, 1):
            c = write_cell(ws2, i, col, val)
            if col == 6:
                c.font = RISK_FONTS.get(worst, DATA_FONT)
                if worst in RISK_FILLS: c.fill = RISK_FILLS[worst]
            if isinstance(val, (int, float)) and val is not None:
                c.alignment = RIGHT
                if isinstance(val, float): c.number_format = "0.0"

    print(f"  Sheet 2: {len(sorted_subs)} substations")

    # ═══════════════════════════════════════════════════════
    # Sheet 3: Manufacturers
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Manufacturers")
    headers3 = ["Manufacturer", "Units", "Avg CHI", "Fault Rate (%)", "Normal (%)",
                "Poor/Critical", "Avg Age (yrs)", "Assessment"]
    widths3 = [30, 8, 10, 12, 10, 12, 10, 18]
    setup_header(ws3, headers3, widths3)

    make_map = defaultdict(lambda: {"n": 0, "chi_s": 0, "chi_n": 0, "faults": 0, "poor": 0, "age_s": 0, "age_n": 0, "normal": 0})
    for t in transformers:
        m = t.get("make") or "Unknown"
        make_map[m]["n"] += 1
        if t.get("chi") is not None: make_map[m]["chi_s"] += t["chi"]; make_map[m]["chi_n"] += 1
        if t.get("fault_label") != "Normal": make_map[m]["faults"] += 1
        else: make_map[m]["normal"] += 1
        if t.get("risk_level") in ("Poor", "Critical"): make_map[m]["poor"] += 1
        if t.get("age_years") is not None and t["age_years"] >= 0: make_map[m]["age_s"] += t["age_years"]; make_map[m]["age_n"] += 1

    for i, (make, d) in enumerate(sorted(make_map.items(), key=lambda x: -x[1]["n"]), 2):
        avg_chi = d["chi_s"] / d["chi_n"] if d["chi_n"] else 0
        fr = d["faults"] / d["n"] * 100
        assess = "Needs review" if d["poor"] > 0 else "High fault rate" if fr > 40 else "Below fleet avg" if avg_chi < 85 else "Satisfactory"
        row = [make, d["n"], round(avg_chi, 1), round(fr, 1), round(d["normal"] / d["n"] * 100, 1),
               d["poor"], round(d["age_s"] / d["age_n"], 1) if d["age_n"] else 0, assess]
        for col, val in enumerate(row, 1):
            c = write_cell(ws3, i, col, val)
            if isinstance(val, (int, float)): c.alignment = RIGHT
            if isinstance(val, float): c.number_format = "0.0"
            if col == 8:
                color = {"Needs review": "C62828", "High fault rate": "E65100", "Below fleet avg": "E65100"}.get(val, "2E7D32")
                c.font = Font(name="Arial", size=9, bold=True, color=color)

    print(f"  Sheet 3: {len(make_map)} manufacturers")

    # ═══════════════════════════════════════════════════════
    # Sheet 4: Action Required (Poor/Critical with Reason)
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Action Required")
    headers4 = ["#", "Equipment No", "Substation", "Make", "kV", "MVA",
                "Age", "CHI", "Risk", "Fault",
                "Reason for Risk Rating",
                "H2", "CH4", "C2H4", "C2H2", "CO", "Moisture",
                "Last Sampled"]
    widths4 = [4, 13, 35, 14, 6, 8, 6, 7, 8, 8, 55, 8, 8, 8, 8, 8, 10, 13]
    setup_header(ws4, headers4, widths4)

    ORANGE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    poor_trs = sorted(
        [t for t in transformers if t.get("risk_level") in ("Poor", "Critical")],
        key=lambda t: t.get("chi") or 999
    )
    for i, t in enumerate(poor_trs, 2):
        gases = t.get("gases", {})
        oil = t.get("oil_quality", {})
        reason = build_risk_reason(t)
        row = [
            i - 1, t.get("equipment_no"),
            t.get("substation_name") or t.get("substation_id") or "",
            t.get("make") or "", t.get("voltage_class"), t.get("capacity_mva"),
            round(t["age_years"], 1) if t.get("age_years") is not None else None,
            round(t["chi"], 1) if t.get("chi") is not None else None,
            t.get("risk_level") or "", t.get("fault_label") or "",
            reason,
            gases.get("H2"), gases.get("CH4"), gases.get("C2H4"),
            gases.get("C2H2"), gases.get("CO"), oil.get("MOISTURE"),
            t.get("latest_sample_date") or "",
        ]
        for col, val in enumerate(row, 1):
            c = write_cell(ws4, i, col, val)
            c.fill = ORANGE_FILL
            if col == 11:  # Reason
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.font = Font(name="Arial", size=9, color="C62828")
            elif col in (8, 9):
                c.font = Font(name="Arial", size=9, bold=True, color="C62828")
            if isinstance(val, (int, float)) and val is not None and col != 11:
                c.alignment = RIGHT
                if isinstance(val, float): c.number_format = "0.0"

    print(f"  Sheet 4: {len(poor_trs)} action-required transformers")

    # ═══════════════════════════════════════════════════════
    # Sheet 5: Model Performance
    # ═══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Model Performance")
    ws5.merge_cells("A1:F1")
    ws5["A1"].value = "Machine Learning Model Performance Summary"
    ws5["A1"].font = TITLE_FONT
    ws5.column_dimensions["A"].width = 18
    for cl in ["B", "C", "D", "E"]: ws5.column_dimensions[cl].width = 14

    r = 3
    fc = models.get("fault_classifier", {})
    ws5.cell(r, 1, "Model 1: Fault Classification").font = Font(name="Arial", size=11, bold=True, color=BLUE)
    r += 1
    for lbl, val in [("Accuracy", f"{fc.get('accuracy',0)*100:.1f}%"), ("Weighted F1", f"{fc.get('weighted_f1',0)*100:.1f}%"),
                      ("Macro F1", f"{fc.get('macro_f1',0)*100:.1f}%"), ("Train/Test", f"{fc.get('n_train',0):,} / {fc.get('n_test',0):,}")]:
        ws5.cell(r, 1, lbl).font = BOLD_FONT; ws5.cell(r, 2, val).font = DATA_FONT; r += 1
    r += 1
    for col, h in enumerate(["Fault Type", "Precision", "Recall", "F1-Score", "Support"], 1):
        c = ws5.cell(r, col, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
    r += 1
    for cls, m in (fc.get("per_class") or {}).items():
        ws5.cell(r, 1, cls).font = BOLD_FONT; ws5.cell(r, 1).border = THIN_BORDER
        for col, val in enumerate([m["precision"]*100, m["recall"]*100, m["f1-score"]*100, int(m["support"])], 2):
            c = ws5.cell(r, col, round(val, 1) if isinstance(val, float) else val)
            c.font = DATA_FONT; c.alignment = RIGHT; c.border = THIN_BORDER
        r += 1

    r += 2
    hi = models.get("health_index", {})
    ws5.cell(r, 1, "Model 2: Health Index Regression").font = Font(name="Arial", size=11, bold=True, color=BLUE); r += 1
    for lbl, val in [("R\u00B2", f"{hi.get('r2',0):.4f}"), ("RMSE", f"{hi.get('rmse',0):.2f}"),
                      ("MAE", f"{hi.get('mae',0):.2f}"), ("Risk Accuracy", f"{hi.get('risk_level_accuracy',0)*100:.1f}%")]:
        ws5.cell(r, 1, lbl).font = BOLD_FONT; ws5.cell(r, 2, val).font = DATA_FONT; r += 1

    r += 2
    fp = models.get("failure_predictor", {})
    ws5.cell(r, 1, "Model 3: Failure Prediction").font = Font(name="Arial", size=11, bold=True, color=BLUE); r += 1
    for lbl, val in [("Accuracy", f"{fp.get('accuracy',0)*100:.1f}%"), ("AUC-ROC", f"{fp.get('auc_roc',0):.3f}"),
                      ("AUC-PR", f"{fp.get('auc_pr',0):.3f}")]:
        ws5.cell(r, 1, lbl).font = BOLD_FONT; ws5.cell(r, 2, val).font = DATA_FONT; r += 1

    print("  Sheet 5: Model performance")

    # ═══════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════
    out = os.path.join(REPORTS_DIR, "AP_Transco_DGA_Fleet_Data_Annexure.xlsx")
    wb.save(out)
    size = os.path.getsize(out) / 1024
    print(f"\n  Saved: {out} ({size:.0f} KB)")


if __name__ == "__main__":
    print("Generating Excel Annexure...")
    main()
    print("Done!")
